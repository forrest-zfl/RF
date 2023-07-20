import openpyxl
from openpyxl.styles import Font,Alignment,Border,Side

# 设置对齐方式
alignment = Alignment(vertical="center",horizontal="center")

# 设置边框的线条
side = Side(color="ff7f50",style="mediumDashed")

# 加载工作薄
wb = openpyxl.load_workbook(r"C:\python07\day17\期末测试成绩.xlsx")

# 获取指定的工作表
sheet = wb.worksheets[0]

# 调整行高和列宽
sheet.row_dimensions[1].height = 30
sheet.column_dimensions["E"].width = 120

sheet['E1'] = "平均分"

# 设置字体
sheet.cell(1,5).font = Font(size=18,bold=True,color='ff1493',name="华文楷体")
# 设置对齐方式
sheet.cell(1,5).alignment = alignment

# 设置单元格边框
sheet.cell(1,5).border = Border(left=side,right=side,top=side,bottom=side)


for i in range(2,8):
# 利用公式计算每个学生的平均分
    sheet[f"E{i}"] = f"=average(B{i}:D{i})"
    sheet.cell(i,5).font = Font(size=12,color='4169e1',italic=True)
    sheet.cell(i,5).alignment = alignment

wb.save("添加样式-期末测试成绩.xlsx")
