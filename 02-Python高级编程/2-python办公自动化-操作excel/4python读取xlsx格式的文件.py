import openpyxl
# 加载一个工作薄
wb = openpyxl.load_workbook(r"C:\python07\day17\笔记\resources\阿里巴巴2020年股票数据.xlsx")

# 获取工作表名称
# print(wb.sheetnames)

# 获取要操作的工作表
sheet = wb.worksheets[0]
#获取工作表中单元格的范围
# print(sheet.dimensions)   # A1:G255

# 获取行数和列数
# print(sheet.max_row,sheet.max_column)  # 255 7

# 获取指定单元格的值
print(sheet.cell(3,3).value)
# print(sheet['C3'].value)
# print(sheet['G255'].value)
# 获取多个单元格
# print(sheet['A2:C5'])
# 获取单元格中所有的数据
for row in range(2,sheet.max_row+1):
    for col in "ABCDEFG":
        value = sheet[f'{col}{row}'].value
        print(value,end="\t")
    print()