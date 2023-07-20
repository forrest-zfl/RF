import random
import openpyxl

# 第一步:创建工作薄
wb = openpyxl.Workbook()

# 第二步:添加工作表
sheet = wb.active
sheet.title = "测试成绩"

# 第三步:写入数据 定义标题
titles = ("姓名","骑马","射箭","摔跤")

# 将标题数据写入到文件中
for col_index,title in enumerate(titles):
    sheet.cell(1,col_index+1,title)

# 定义学生姓名
students = ("欣迪","中文","文定","珊珊","惠君","俊德")

# 将姓名和成绩写入到文件中去
for row_index,student in enumerate(students):
    sheet.cell(row_index+2,1,student)
    for col_index in range(2,5):
        sheet.cell(row_index+2,col_index,random.randrange(50,101))

# 将数据保存到工作簿中去
wb.save("期末测试成绩.xlsx")

# 注意:在 2007年之后的xlsx文件中的行和列的索引是从1开始   
#      在2007年之前的xls文件中的行和列的索引是从0开始